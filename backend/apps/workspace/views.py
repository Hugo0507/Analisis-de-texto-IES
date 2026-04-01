"""
Workspace Views

API para el Laboratorio: subida de PDFs e inferencia con modelos del corpus.

Arquitectura de procesos:
    - La inferencia corre en un subprocess completamente separado via
      `python manage.py run_inference <workspace_id>`. Esto evita la
      heap corruption causada por fork() heredando estado de numpy/scipy
      del proceso gunicorn padre.
    - Un Thread monitor vigila el subprocess; si crashea o excede el timeout
      de 5 minutos, marca el workspace como error automáticamente.
"""

import io
import json
import logging
import subprocess
import sys
import threading

from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Workspace, WorkspaceDocument
from .serializers import (
    WorkspaceSerializer,
    WorkspaceCreateSerializer,
    WorkspaceUploadSerializer,
)

# Parámetros por defecto para inferencia
DEFAULT_INFERENCE_PARAMS = {
    'num_top_terms': 50,
    'strip_references': True,
    'min_word_length': 2,
    'ner_entity_types': [],   # vacío = heredar del NER de referencia
}

logger = logging.getLogger(__name__)

# Timeout máximo para el proceso de inferencia (5 minutos)
INFERENCE_TIMEOUT_SECONDS = 5 * 60


# ── Endpoints ────────────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def workspace_list(request):
    """Listar workspaces del usuario o crear uno nuevo."""
    if request.method == 'GET':
        workspaces = Workspace.objects.filter(created_by=request.user).prefetch_related('documents')
        serializer = WorkspaceSerializer(workspaces, many=True)
        return Response(serializer.data)

    serializer = WorkspaceCreateSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        workspace = serializer.save()
        return Response(WorkspaceSerializer(workspace).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def workspace_detail(request, workspace_id):
    """Obtener o eliminar un workspace."""
    try:
        workspace = Workspace.objects.prefetch_related('documents').get(
            id=workspace_id,
            created_by=request.user
        )
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        workspace.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    return Response(WorkspaceSerializer(workspace).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def workspace_upload(request, workspace_id):
    """Subir un PDF al workspace."""
    try:
        workspace = Workspace.objects.get(id=workspace_id, created_by=request.user)
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status == Workspace.STATUS_PROCESSING:
        return Response(
            {'error': 'El workspace está procesando. Espera a que termine antes de subir más archivos.'},
            status=status.HTTP_409_CONFLICT
        )

    serializer = WorkspaceUploadSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    uploaded_file = serializer.validated_data['file']
    doc = WorkspaceDocument.objects.create(
        workspace=workspace,
        file=uploaded_file,
        original_filename=uploaded_file.name,
        file_size=uploaded_file.size,
        status=WorkspaceDocument.STATUS_PENDING,
    )

    return Response(
        {
            'id': doc.id,
            'original_filename': doc.original_filename,
            'file_size': doc.file_size,
            'status': doc.status,
        },
        status=status.HTTP_201_CREATED
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def workspace_run(request, workspace_id):
    """
    Iniciar inferencia sobre los documentos subidos al workspace.

    Lanza un subprocess completamente separado (python manage.py run_inference)
    y un Thread monitor que detecta crashes o timeouts.
    """
    try:
        workspace = Workspace.objects.prefetch_related('documents').get(
            id=workspace_id,
            created_by=request.user
        )
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status == Workspace.STATUS_PROCESSING:
        return Response(
            {'error': 'El workspace ya está procesando.'},
            status=status.HTTP_409_CONFLICT
        )

    pending_docs = workspace.documents.filter(
        status__in=[WorkspaceDocument.STATUS_PENDING, WorkspaceDocument.STATUS_READY]
    )
    if not pending_docs.exists():
        return Response(
            {'error': 'No hay documentos para procesar. Sube al menos un PDF.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    workspace.status = Workspace.STATUS_PROCESSING
    workspace.progress_percentage = 0
    workspace.error_message = None
    workspace.results = {}
    workspace.save()

    # Lanzar subprocess completamente aislado (sin fork, sin heap compartido)
    process = subprocess.Popen(
        [sys.executable, 'manage.py', 'run_inference', str(workspace.id)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )

    logger.info(
        f"[WS {workspace_id}] Subprocess de inferencia lanzado (PID: {process.pid})"
    )

    # Monitor: vigila el subprocess y actualiza status si crashea/timeout
    monitor = threading.Thread(
        target=_monitor_inference_subprocess,
        args=(process, str(workspace.id)),
        daemon=True,
    )
    monitor.start()

    return Response(
        {'status': 'processing', 'workspace_id': str(workspace.id)},
        status=status.HTTP_202_ACCEPTED
    )


# ── Monitor del subprocess de inferencia ─────────────────────────────────

def _monitor_inference_subprocess(process: subprocess.Popen, workspace_id: str):
    """
    Vigila el subprocess de inferencia.

    Si el proceso termina con returncode != 0 (crash, OOM, etc.)
    o excede el timeout, marca el workspace como error.
    Esto garantiza que NUNCA quede un workspace atascado en 'processing'.
    """
    try:
        try:
            stdout, stderr = process.communicate(timeout=INFERENCE_TIMEOUT_SECONDS)
        except subprocess.TimeoutExpired:
            logger.error(
                f"[WS {workspace_id}] Timeout de {INFERENCE_TIMEOUT_SECONDS}s "
                f"alcanzado. Terminando subprocess..."
            )
            process.kill()
            stdout, stderr = process.communicate(timeout=10)

            _mark_workspace_error(
                workspace_id,
                f"La inferencia excedió el tiempo máximo de "
                f"{INFERENCE_TIMEOUT_SECONDS // 60} minutos y fue cancelada."
            )
            return

        if process.returncode != 0:
            stderr_text = stderr.decode('utf-8', errors='replace')[-500:] if stderr else ''
            logger.error(
                f"[WS {workspace_id}] Subprocess terminó con código {process.returncode}. "
                f"stderr: {stderr_text}"
            )
            _mark_workspace_error(
                workspace_id,
                f"El proceso de inferencia terminó con error "
                f"(código: {process.returncode}). Intenta de nuevo."
            )
            return

        logger.info(f"[WS {workspace_id}] Subprocess de inferencia finalizó OK")

    except Exception as e:
        logger.exception(f"[WS {workspace_id}] Error en monitor: {e}")
        _mark_workspace_error(workspace_id, f"Error interno del monitor: {e}")


def _mark_workspace_error(workspace_id: str, message: str):
    """Marca un workspace como error (solo si aún está en processing)."""
    try:
        workspace = Workspace.objects.get(id=workspace_id)
        if workspace.status == Workspace.STATUS_PROCESSING:
            workspace.status = Workspace.STATUS_ERROR
            workspace.error_message = message[:500]
            workspace.save()
            logger.info(f"[WS {workspace_id}] Marcado como error: {message}")
    except Exception as e:
        logger.error(f"[WS {workspace_id}] No se pudo marcar error: {e}")


# ── Stopwords ─────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workspace_corpus_stopwords(request):
    """
    Obtener las stopwords del corpus para un dataset dado (sin workspace previo).

    Query params:
        dataset_id (int): ID del dataset de referencia.
    """
    dataset_id_raw = request.query_params.get('dataset_id')
    if not dataset_id_raw:
        return Response({'error': 'dataset_id requerido.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        dataset_id = int(dataset_id_raw)
    except ValueError:
        return Response({'error': 'dataset_id debe ser un entero.'}, status=status.HTTP_400_BAD_REQUEST)

    from apps.workspace.inference import get_inference_stopwords
    corpus_stopwords = sorted(get_inference_stopwords(dataset_id=dataset_id))
    return Response({'corpus_stopwords': corpus_stopwords})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workspace_stopwords(request, workspace_id):
    """
    Obtener las stopwords activas del workspace.

    Retorna:
    - corpus_stopwords: palabras del corpus (solo lectura)
    - custom_stopwords: palabras personalizadas del usuario para esta sesión
    """
    try:
        workspace = Workspace.objects.get(id=workspace_id, created_by=request.user)
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    from apps.workspace.inference import get_inference_stopwords
    corpus_stopwords = sorted(get_inference_stopwords(dataset_id=workspace.dataset_id))

    return Response({
        'corpus_stopwords': corpus_stopwords,
        'custom_stopwords': workspace.custom_stopwords or [],
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def workspace_stopwords_update(request, workspace_id):
    """
    Añadir o quitar stopwords personalizadas.

    Body: { "add": ["word1", "word2"], "remove": ["word3"] }
    """
    try:
        workspace = Workspace.objects.get(id=workspace_id, created_by=request.user)
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status == Workspace.STATUS_PROCESSING:
        return Response(
            {'error': 'No se puede modificar stopwords mientras la inferencia está en curso.'},
            status=status.HTTP_409_CONFLICT
        )

    to_add = request.data.get('add', [])
    to_remove = request.data.get('remove', [])

    if not isinstance(to_add, list) or not isinstance(to_remove, list):
        return Response(
            {'error': 'Los campos "add" y "remove" deben ser listas.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    current = set(workspace.custom_stopwords or [])

    # Normalizar: lowercase, strip, solo palabras no vacías
    new_words = {w.strip().lower() for w in to_add if isinstance(w, str) and w.strip()}
    remove_words = {w.strip().lower() for w in to_remove if isinstance(w, str) and w.strip()}

    current = (current | new_words) - remove_words
    workspace.custom_stopwords = sorted(current)
    workspace.save(update_fields=['custom_stopwords'])

    return Response({
        'custom_stopwords': workspace.custom_stopwords,
        'added': sorted(new_words - remove_words),
        'removed': sorted(remove_words & (current | remove_words)),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def workspace_stopwords_import(request, workspace_id):
    """
    Importar stopwords desde un archivo TXT (una palabra por línea).

    Las palabras importadas se añaden a las custom_stopwords existentes.
    """
    try:
        workspace = Workspace.objects.get(id=workspace_id, created_by=request.user)
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status == Workspace.STATUS_PROCESSING:
        return Response(
            {'error': 'No se puede importar stopwords mientras la inferencia está en curso.'},
            status=status.HTTP_409_CONFLICT
        )

    uploaded = request.FILES.get('file')
    if not uploaded:
        return Response({'error': 'No se recibió ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)

    allowed_extensions = {'.txt', '.csv'}
    import os
    ext = os.path.splitext(uploaded.name)[1].lower()
    if ext not in allowed_extensions:
        return Response(
            {'error': f'Solo se aceptan archivos .txt o .csv. Recibido: {uploaded.name}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        content = uploaded.read().decode('utf-8', errors='replace')
    except Exception:
        return Response({'error': 'No se pudo leer el archivo.'}, status=status.HTTP_400_BAD_REQUEST)

    # Parsear: una palabra por línea (ignorar líneas vacías y comentarios #)
    imported = set()
    for line in content.splitlines():
        word = line.strip().lower()
        if word and not word.startswith('#'):
            imported.add(word)

    if not imported:
        return Response({'error': 'El archivo no contiene palabras válidas.'}, status=status.HTTP_400_BAD_REQUEST)

    current = set(workspace.custom_stopwords or [])
    current |= imported
    workspace.custom_stopwords = sorted(current)
    workspace.save(update_fields=['custom_stopwords'])

    return Response({
        'custom_stopwords': workspace.custom_stopwords,
        'imported_count': len(imported),
        'imported_words': sorted(imported),
    })


# ── Import ───────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def workspace_import_config(request):
    """
    Crear un workspace a partir de una configuración exportada.

    Body JSON:
        dataset_id  (int)  — Dataset de referencia para el nuevo workspace.
        config      (dict) — JSON exportado con workspace_export_config.

    Devuelve:
        workspace_id  (str)  — UUID del nuevo workspace creado.
        has_results   (bool) — True si el config incluía resultados previos
                               y el workspace fue creado en estado 'completed'.
        warnings      (list) — Mensajes por modelos que ya no existen en DB.
    """
    dataset_id = request.data.get('dataset_id')
    config = request.data.get('config')

    if not dataset_id:
        return Response({'error': 'dataset_id es requerido.'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(config, dict):
        return Response({'error': 'config debe ser un objeto JSON válido.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validar que el dataset existe
    from apps.datasets.models import Dataset
    try:
        dataset = Dataset.objects.get(id=dataset_id)
    except Dataset.DoesNotExist:
        return Response({'error': f'Dataset {dataset_id} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    # Extraer modelos del config
    models_cfg = config.get('models') or {}

    warnings = []
    resolved = {
        'bow_id': None,
        'tfidf_id': None,
        'topic_model_id': None,
        'ner_id': None,
        'bertopic_id': None,
    }

    # Validar existencia de cada modelo en DB
    _MODEL_CHECKS = [
        ('bow', 'bow_id', 'apps.bag_of_words.models', 'BagOfWords', 'BoW'),
        ('tfidf', 'tfidf_id', 'apps.tfidf_analysis.models', 'TfIdfAnalysis', 'TF-IDF'),
        ('topic_model', 'topic_model_id', 'apps.topic_modeling.models', 'TopicModeling', 'Topic Model'),
        ('ner', 'ner_id', 'apps.ner_analysis.models', 'NerAnalysis', 'NER'),
        ('bertopic', 'bertopic_id', 'apps.bertopic.models', 'BERTopicAnalysis', 'BERTopic'),
    ]

    for cfg_key, field, module_path, class_name, label in _MODEL_CHECKS:
        entry = models_cfg.get(cfg_key)
        if not entry or entry.get('id') is None:
            continue  # modelo no seleccionado en el config original → se omite
        model_id = entry['id']
        try:
            import importlib
            mod = importlib.import_module(module_path)
            cls = getattr(mod, class_name)
            cls.objects.get(id=model_id)
            resolved[field] = model_id
        except Exception:
            model_name = entry.get('name') or f'#{model_id}'
            warnings.append(f'{label} "{model_name}" (id={model_id}) ya no existe en la DB.')

    # Determinar si el config trae resultados previos
    results_data = config.get('results') or {}
    has_results = bool(results_data)

    # Crear workspace
    ws_kwargs = {
        'created_by': request.user,
        'dataset': dataset,
        'custom_stopwords': config.get('custom_stopwords') or [],
        'inference_params': config.get('inference_params') or {},
        **resolved,
    }

    if has_results:
        ws_kwargs['results'] = results_data
        ws_kwargs['status'] = Workspace.STATUS_COMPLETED
        ws_kwargs['progress_percentage'] = 100

    workspace = Workspace.objects.create(**ws_kwargs)

    logger.info(
        f"[WS {workspace.id}] Creado por import_config "
        f"(dataset={dataset_id}, has_results={has_results}, warnings={len(warnings)})"
    )

    return Response({
        'workspace_id': str(workspace.id),
        'has_results': has_results,
        'warnings': warnings,
    }, status=status.HTTP_201_CREATED)


# ── Export ────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workspace_export_excel(request, workspace_id):
    """
    Exportar resultados del workspace como archivo Excel (.xlsx).

    Hojas generadas:
      1. Metadatos  — configuración de la sesión
      2. BoW        — top términos (si se ejecutó)
      3. TF-IDF     — top términos (si se ejecutó)
      4. Temas      — distribución + afinidad (si se ejecutó)
      5. NER        — distribución + top entidades (si se ejecutó)
      6. BERTopic   — distribución + asignaciones (si se ejecutó)
    """
    try:
        workspace = Workspace.objects.select_related('dataset').get(
            id=workspace_id, created_by=request.user
        )
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status != Workspace.STATUS_COMPLETED:
        return Response(
            {'error': 'El workspace aún no está completado.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        xlsx_bytes = _build_excel_bytes(workspace)
    except ImportError:
        return Response(
            {'error': 'openpyxl no está instalado en este entorno.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    except Exception as exc:
        logger.exception(f"[WS {workspace_id}] Error generando Excel: {exc}")
        return Response(
            {'error': f'Error generando el Excel: {exc}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    date_str = workspace.created_at.strftime('%Y%m%d')
    dataset_slug = workspace.dataset.name[:30].replace(' ', '_')
    filename = f"lab_{dataset_slug}_{date_str}.xlsx"

    response = HttpResponse(
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workspace_export_config(request, workspace_id):
    """
    Exportar configuración + resultados del workspace como JSON.

    El JSON incluye modelos seleccionados, stopwords, parámetros y
    resultados completos. Puede reimportarse para recrear la sesión.
    """
    try:
        workspace = Workspace.objects.select_related('dataset').get(
            id=workspace_id, created_by=request.user
        )
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if workspace.status != Workspace.STATUS_COMPLETED:
        return Response(
            {'error': 'El workspace aún no está completado.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    config = _build_config_dict(workspace)
    json_bytes = json.dumps(config, ensure_ascii=False, indent=2).encode('utf-8')

    date_str = workspace.created_at.strftime('%Y%m%d')
    dataset_slug = workspace.dataset.name[:30].replace(' ', '_')
    filename = f"lab_config_{dataset_slug}_{date_str}.json"

    response = HttpResponse(json_bytes, content_type='application/json; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Export helpers ─────────────────────────────────────────────────────────────

def _build_excel_bytes(workspace) -> bytes:
    """Construye el workbook Excel y devuelve los bytes crudos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    results = workspace.results or {}

    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(fill_type='solid', fgColor='334155')
    TITLE_FONT = Font(bold=True, size=11)

    def _hdr(sheet):
        """Aplica estilo de cabecera a la última fila escrita."""
        for cell in sheet[sheet.max_row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    wb = Workbook()

    # ── Hoja 1: Metadatos ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Metadatos'

    params = workspace.inference_params or {}
    r = results

    ws.append(['Campo', 'Valor'])
    _hdr(ws)

    meta = [
        ('Workspace ID',        str(workspace.id)),
        ('Dataset',             workspace.dataset.name),
        ('Fecha análisis',      workspace.created_at.strftime('%Y-%m-%d %H:%M UTC')),
        ('Documentos procesados', r.get('document_count', 0)),
        ('Modelo BoW',          (r.get('bow') or {}).get('reference_bow_name', 'N/A')),
        ('Modelo TF-IDF',       (r.get('tfidf') or {}).get('reference_tfidf_name', 'N/A')),
        ('Modelo Topics',       (r.get('topics') or {}).get('reference_topic_model_name', 'N/A')),
        ('Modelo NER',          (r.get('ner') or {}).get('reference_ner_name', 'N/A')),
        ('Modelo BERTopic',     (r.get('bertopic') or {}).get('reference_bertopic_name', 'N/A')),
        ('Top términos',        params.get('num_top_terms', 50)),
        ('Long. mín. token',    params.get('min_word_length', 2)),
        ('Cortar referencias',  'Sí' if params.get('strip_references', True) else 'No'),
        ('Tipos NER',           ', '.join(params.get('ner_entity_types', [])) or 'Heredado del análisis'),
        ('Stopwords propias',   ', '.join(workspace.custom_stopwords or []) or 'Ninguna'),
    ]
    for label, value in meta:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55

    # ── Hoja 2: BoW ─────────────────────────────────────────────────────
    bow = results.get('bow') or {}
    if bow and not bow.get('error') and bow.get('top_terms'):
        ws2 = wb.create_sheet('BoW — Top Términos')
        ws2.append(['#', 'Término', 'Frecuencia'])
        _hdr(ws2)
        for t in bow['top_terms']:
            ws2.append([t.get('rank', ''), t.get('term', ''), int(t.get('score', 0))])
        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 28
        ws2.column_dimensions['C'].width = 14

    # ── Hoja 3: TF-IDF ──────────────────────────────────────────────────
    tfidf = results.get('tfidf') or {}
    if tfidf and not tfidf.get('error') and tfidf.get('top_terms'):
        ws3 = wb.create_sheet('TF-IDF — Top Términos')
        ws3.append(['#', 'Término', 'Peso TF-IDF'])
        _hdr(ws3)
        for t in tfidf['top_terms']:
            ws3.append([t.get('rank', ''), t.get('term', ''), round(float(t.get('score', 0)), 4)])
        ws3.column_dimensions['A'].width = 5
        ws3.column_dimensions['B'].width = 28
        ws3.column_dimensions['C'].width = 14

    # ── Hoja 4: Temas ────────────────────────────────────────────────────
    topics = results.get('topics') or {}
    if topics and not topics.get('error'):
        ws4 = wb.create_sheet('Temas')

        ws4.append(['Distribución de temas en documentos nuevos'])
        ws4.cell(row=ws4.max_row, column=1).font = TITLE_FONT
        ws4.append(['Tema', 'N° Documentos', '%'])
        _hdr(ws4)
        for t in (topics.get('topic_distribution') or []):
            ws4.append([
                t.get('topic_label', f"Tema {t.get('topic_id')}"),
                t.get('document_count', 0),
                t.get('percentage', 0),
            ])

        if topics.get('all_topics_affinity'):
            ws4.append([])
            ws4.append(['Afinidad promedio con todos los temas del corpus'])
            ws4.cell(row=ws4.max_row, column=1).font = TITLE_FONT
            ws4.append(['Tema', 'Peso', '%', 'Top palabras'])
            _hdr(ws4)
            for a in topics['all_topics_affinity']:
                words = a.get('top_words', [])
                words_str = ', '.join(
                    w.get('word', str(w)) if isinstance(w, dict) else str(w)
                    for w in words[:5]
                )
                ws4.append([
                    a.get('topic_label', f"Tema {a.get('topic_id')}"),
                    round(float(a.get('weight', 0)), 4),
                    a.get('percentage', 0),
                    words_str,
                ])

        ws4.column_dimensions['A'].width = 32
        ws4.column_dimensions['B'].width = 14
        ws4.column_dimensions['C'].width = 8
        ws4.column_dimensions['D'].width = 42

    # ── Hoja 5: NER ──────────────────────────────────────────────────────
    ner = results.get('ner') or {}
    if ner and not ner.get('error'):
        ws5 = wb.create_sheet('NER — Entidades')

        ws5.append(['Distribución por tipo de entidad'])
        ws5.cell(row=ws5.max_row, column=1).font = TITLE_FONT
        ws5.append(['Tipo', 'Total ocurrencias', 'Entidades únicas', '%'])
        _hdr(ws5)
        for item in (ner.get('entity_distribution') or []):
            ws5.append([
                item.get('type', ''),
                item.get('count', 0),
                item.get('unique_entities', 0),
                round(float(item.get('percentage', 0)), 1),
            ])

        top_by_type = ner.get('top_entities_by_type') or {}
        if top_by_type:
            ws5.append([])
            ws5.append(['Top entidades por tipo'])
            ws5.cell(row=ws5.max_row, column=1).font = TITLE_FONT
            ws5.append(['Tipo', 'Entidad', 'Frecuencia'])
            _hdr(ws5)
            for etype, entities in top_by_type.items():
                for e in (entities or []):
                    ws5.append([etype, e.get('text', ''), e.get('count', 0)])

        ws5.column_dimensions['A'].width = 16
        ws5.column_dimensions['B'].width = 32
        ws5.column_dimensions['C'].width = 18
        ws5.column_dimensions['D'].width = 8

    # ── Hoja 6: BERTopic ─────────────────────────────────────────────────
    bert = results.get('bertopic') or {}
    if bert and not bert.get('error'):
        ws6 = wb.create_sheet('BERTopic — Similitud')

        method_note = bert.get('method_note', '')
        ws6.append([f"Método: {bert.get('method', 'keyword_similarity')}  —  {method_note}"])
        ws6.cell(row=1, column=1).font = Font(italic=True, color='888888')
        ws6.append([])

        ws6.append(['Distribución de documentos por tema'])
        ws6.cell(row=ws6.max_row, column=1).font = TITLE_FONT
        ws6.append(['Tema', 'N° Documentos', '%'])
        _hdr(ws6)
        for t in (bert.get('topic_distribution') or []):
            ws6.append([
                t.get('topic_label', f"Tema {t.get('topic_id')}"),
                t.get('document_count', 0),
                t.get('percentage', 0),
            ])

        assignments = bert.get('document_assignments') or []
        if assignments:
            ws6.append([])
            ws6.append(['Asignación por documento'])
            ws6.cell(row=ws6.max_row, column=1).font = TITLE_FONT
            ws6.append(['# Doc', 'Tema dominante', 'Similitud (%)'])
            _hdr(ws6)
            for da in assignments:
                ws6.append([
                    da.get('document_index', 0) + 1,
                    da.get('dominant_topic_label', ''),
                    round(float(da.get('similarity_score', 0)) * 100, 1),
                ])

        ws6.column_dimensions['A'].width = 8
        ws6.column_dimensions['B'].width = 35
        ws6.column_dimensions['C'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_config_dict(workspace) -> dict:
    """Construye el diccionario de configuración exportable."""
    results = workspace.results or {}
    bow_r = results.get('bow') or {}
    tfidf_r = results.get('tfidf') or {}
    topics_r = results.get('topics') or {}
    ner_r = results.get('ner') or {}
    bert_r = results.get('bertopic') or {}

    def _model_entry(ws_id, name_key, result_dict):
        if ws_id is None:
            return None
        return {'id': ws_id, 'name': result_dict.get(name_key)}

    return {
        'schema_version': '1.0',
        'created_at': workspace.created_at.isoformat(),
        'workspace_id': str(workspace.id),
        'dataset_name': workspace.dataset.name,
        'models': {
            'bow': _model_entry(workspace.bow_id, 'reference_bow_name', bow_r),
            'tfidf': _model_entry(workspace.tfidf_id, 'reference_tfidf_name', tfidf_r),
            'topic_model': _model_entry(workspace.topic_model_id, 'reference_topic_model_name', topics_r),
            'ner': _model_entry(workspace.ner_id, 'reference_ner_name', ner_r),
            'bertopic': _model_entry(workspace.bertopic_id, 'reference_bertopic_name', bert_r),
        },
        'custom_stopwords': workspace.custom_stopwords or [],
        'inference_params': workspace.inference_params or {},
        'results': results,
    }
